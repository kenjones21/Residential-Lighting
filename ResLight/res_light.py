import os
from openpyxl import load_workbook
from operator import itemgetter, attrgetter
class ResLight:
    def __init__(self, full_filename, source):
        if source is "DOE":
            filename, extension = os.path.splitext(full_filename)
            if extension != ".xlsx":
                raise Exception("DOE files must be in xlsx format")
            print("Loading workbook...", end="", flush=True)
            wb = load_workbook(full_filename)
            print("Done")
            tool = wb.get_sheet_by_name("C) Tool")
            sort_order = [tool["A2"].value, tool["B2"].value, tool["C2"].value, tool["D2"].value]
            self.light_data = LightData(sort_order)
            for row in tool.iter_rows(min_row = 3, max_row = 16197): #16197
                attributes = {}
                for cell in row:
                    col = cell.column
                    if col is "A" or col is "B" or col is "C" or col is "D":
                        category = tool[col + "2"].value
                        attributes[category] = str(cell.value)
                    else:
                        category = tool[col + "2"].value
                        attributes[category] = cell.value
                light_datum_temp = LightDatum(attributes)
                self.light_data.insert(light_datum_temp)
            print(len(self.light_data))
            self.light_data.attributes = sorted(self.light_data.attributes,
                                     key = itemgetter(sort_order[0], sort_order[1],
                                                      sort_order[2], sort_order[3]))
        else:
            raise Exception("Only DOE files accepted at this time")

class LightDatum:
    def __init__(self, attributes):
        self.attributes = attributes

    def compare_to(self, other, sort_attributes):
        for attr in sort_attributes:
            if self[attr] != other[attr]:
                return self[attr] < other[attr]
        return True
            
class LightData:
    def __init__(self, hierarchy):
        "hierarchy: list of strings indicating hierarchy levels"
        self.hierarchy = hierarchy
        self.data = Category("Master")
    
    def insert(self, light_datum):
        current_category = self.data
        for element in self.hierarchy:
            cat_name = light_datum.attributes[element]
            if current_category.has_child(cat_name):
                current_category = current_category.get_child(cat_name)
            else:
                new_cat = Category(cat_name)
                current_category.add(new_cat)
        current_category.add_datum(light_datum)

    def insertion_index(self, light_datum):
        lbound = 0
        rbound = len(self.data)
        if lbound == rbound:
            return lbound

class Category:
    def __init__(self, name):
        self.name = name
        self.children = []

    def add(self, element):
        "Adds a child category to this category"
        self.children.append(element)

    def has_child(self, name):
        for child in self.children:
            if child.name == name:
                return True
        return False

    def get_child(self, name):
        for child in self.children:
            if child.name == name:
                return child

    def add_datum(self, light_datum):
        self.light_datum = light_datum

r = ResLight('../Data/reslight_DOE_2012.xlsx', 'DOE')
        
